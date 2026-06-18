# -*- coding: utf-8 -*-
import pandas as pd, json
from collections import defaultdict
from datetime import datetime
from processador import normalizar_sku
from cruzamento import criar_indices

ARQ = 'RelatorioItensPorVendedor_3280c9c9-62bc-40c7-aa89-fd0958eeb05b.xlsx'
df = pd.read_excel(ARQ, dtype=str)
c = list(df.columns)
C_CODVEND, C_VEND, C_CODPROD, C_DTCAP, C_CODREV, C_REV, C_CANAL, C_CODPED, C_NF = \
    c[0], c[1], c[2], c[9], c[13], c[14], c[27], c[8], c[17]


def cl(s):
    if s is None:
        return ''
    s = str(s).strip()
    return '' if s in ('nan', 'None', '<NA>') else s


for col in [C_CODVEND, C_VEND, C_CODPROD, C_DTCAP, C_CODREV, C_REV, C_CANAL, C_CODPED, C_NF]:
    df[col] = df[col].map(cl)
df[C_CODPED] = df[C_CODPED].replace('', pd.NA).ffill().fillna('')
df['PEDIDO'] = df.apply(lambda r: r[C_CODPED] or r[C_NF], axis=1)

ip, _ = criar_indices('produtos.db')
mc = json.load(open('marcas_catalog.json', encoding='utf-8'))


def marca_de(sku):
    sn = normalizar_sku(sku)
    e = ip.get(sn)
    if e and e.get('marca'):
        return e['marca']
    if sn in mc and mc[sn].get('marca'):
        return mc[sn]['marca']
    return ''


def disp_marca(m):
    return 'O Boticário' if m.startswith('oBotic') else m


df['MARCA'] = df[C_CODPROD].map(marca_de)
df['UNID'] = df[C_CANAL].map(
    lambda x: 'Matriz Penedo' if x.startswith('13707')
    else ('Filial Palmeira dos Índios' if x.startswith('13706') else 'Outra'))


def pdt(s):
    for f in ['%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            pass
    return None


df['DT'] = df[C_DTCAP].map(pdt)

# pedido -> atributos
ped = defaultdict(lambda: {'rev': '', 'vend': '', 'vnome': '', 'unid': '', 'dt': None, 'marcas': set()})
for _, r in df.iterrows():
    p = r['PEDIDO']
    if not p:
        continue
    d = ped[p]
    d['rev'] = d['rev'] or r[C_CODREV]
    if not d['vend'] and r[C_CODVEND]:
        d['vend'] = r[C_CODVEND]
        d['vnome'] = r[C_VEND]
    d['unid'] = d['unid'] or r['UNID']
    if d['dt'] is None or (r['DT'] and r['DT'] < d['dt']):
        d['dt'] = r['DT']
    if r['MARCA']:
        d['marcas'].add(r['MARCA'])

# cliente -> pedidos
cli = defaultdict(list)
rev_nome = {}
for _, r in df.iterrows():
    if r[C_CODREV] and r[C_CODREV] not in rev_nome:
        rev_nome[r[C_CODREV]] = r[C_REV]
for p, d in ped.items():
    if d['rev']:
        cli[d['rev']].append(d)

BONUS = 0.02
META = 72.0
detalhe = []
agg = defaultdict(lambda: {'vnome': '', 'unid': '', 'ativos': 0, 'multi': 0,
                           'first': 0, 'built': 0, 'mono': 0, 'pontos': 0.0})
for rev, pedidos in cli.items():
    pedidos = [p for p in pedidos if p['dt']]
    if not pedidos:
        continue
    pedidos.sort(key=lambda x: x['dt'])
    first = pedidos[0]
    v = first['vend']
    vchave = v if v else 'SEM_VEND'
    vnome = first['vnome'] if v else '(sem vendedor)'
    unid = first['unid']
    marcas_tot = set()
    for p in pedidos:
        marcas_tot |= p['marcas']
    is_multi = len(marcas_tot) >= 2
    first_multi = len(first['marcas']) >= 2
    peso = 0.0
    if is_multi:
        peso = (1.0 + BONUS) if first_multi else 1.0
    m = agg[vchave]
    m['vnome'] = vnome
    m['unid'] = unid
    m['ativos'] += 1
    m['pontos'] += peso
    if is_multi:
        m['multi'] += 1
        if first_multi:
            m['first'] += 1
        else:
            m['built'] += 1
    else:
        m['mono'] += 1
    detalhe.append({
        'Unidade': unid, 'Vendedor (1º pedido)': vnome, 'Cód. Revendedor': rev,
        'Revendedor': rev_nome.get(rev, ''), 'Nº Pedidos': len(pedidos),
        'Marcas Distintas': len(marcas_tot),
        'Marcas': ', '.join(sorted(disp_marca(x) for x in marcas_tot)),
        'Multimarca?': 'Sim' if is_multi else 'Não',
        '1º Pedido já Multimarca?': 'Sim' if (is_multi and first_multi) else ('Não' if is_multi else '—'),
        'Pontos': round(peso, 2),
    })

# resumo
resumo = []
for vchave, m in agg.items():
    pct = min(100.0, m['pontos'] / m['ativos'] * 100) if m['ativos'] else 0
    resumo.append({
        'Unidade': m['unid'], 'Vendedor': m['vnome'], 'Código': ('' if vchave == 'SEM_VEND' else vchave),
        'Clientes Ativos': m['ativos'], 'Clientes Multimarca': m['multi'],
        '→ já no 1º pedido': m['first'], '→ construídos no ciclo': m['built'],
        'Clientes Monomarca': m['mono'],
        'Pontos (c/ bônus +2%)': round(m['pontos'], 2),
        '% Multimarca (teto 100%)': round(pct, 1),
        'Meta': META, 'Atingiu Meta?': 'Sim' if pct >= META else 'Não',
    })

# escrever xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HFILL = PatternFill('solid', fgColor='305496')
HFONT = Font(bold=True, color='FFFFFF')
OKF = PatternFill('solid', fgColor='C6EFCE')
NOF = PatternFill('solid', fgColor='FFC7CE')
TOTF = PatternFill('solid', fgColor='D9E1F2')
BOLD = Font(bold=True)


def style_header(ws, ncol):
    for j in range(1, ncol + 1):
        cc = ws.cell(row=1, column=j)
        cc.fill = HFILL
        cc.font = HFONT
        cc.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 42
    ws.auto_filter.ref = f'A1:{get_column_letter(ncol)}{ws.max_row}'


# Aba Resumo
ws = wb.active
ws.title = 'Resumo por Vendedor'
cols1 = ['Unidade', 'Vendedor', 'Código', 'Clientes Ativos', 'Clientes Multimarca',
         '→ já no 1º pedido', '→ construídos no ciclo', 'Clientes Monomarca',
         'Pontos (c/ bônus +2%)', '% Multimarca (teto 100%)', 'Meta', 'Atingiu Meta?']
ws.append(cols1)
ordem_un = ['Matriz Penedo', 'Filial Palmeira dos Índios']
un_tot = defaultdict(lambda: [0, 0.0])
linhas_ord = sorted(resumo, key=lambda x: (ordem_un.index(x['Unidade']) if x['Unidade'] in ordem_un else 9,
                                           -x['% Multimarca (teto 100%)']))
for row in linhas_ord:
    ws.append([row[k] for k in cols1])
    r = ws.max_row
    ws.cell(row=r, column=10).number_format = '0.0"%"'
    ws.cell(row=r, column=11).number_format = '0"%"'
    ws.cell(row=r, column=12).fill = OKF if row['Atingiu Meta?'] == 'Sim' else NOF
    un_tot[row['Unidade']][0] += row['Clientes Ativos']
    un_tot[row['Unidade']][1] += row['Pontos (c/ bônus +2%)']
for un in ordem_un:
    a, p = un_tot[un]
    if a:
        pct = min(100.0, p / a * 100)
        ws.append([f'TOTAL {un}', '', '', a, '', '', '', '', round(p, 2), round(pct, 1),
                   META, 'Sim' if pct >= META else 'Não'])
        r = ws.max_row
        for j in range(1, 13):
            ws.cell(row=r, column=j).fill = TOTF
            ws.cell(row=r, column=j).font = BOLD
        ws.cell(row=r, column=10).number_format = '0.0"%"'
        ws.cell(row=r, column=11).number_format = '0"%"'
widths1 = [26, 38, 12, 14, 16, 15, 18, 16, 18, 18, 8, 13]
for i, w in enumerate(widths1, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(cols1))

# Aba Detalhe Clientes
ws2 = wb.create_sheet('Detalhe Clientes')
cols2 = ['Unidade', 'Vendedor (1º pedido)', 'Cód. Revendedor', 'Revendedor', 'Nº Pedidos',
         'Marcas Distintas', 'Marcas', 'Multimarca?', '1º Pedido já Multimarca?', 'Pontos']
ws2.append(cols2)
detalhe.sort(key=lambda x: (x['Unidade'], x['Vendedor (1º pedido)'], -x['Marcas Distintas']))
for d in detalhe:
    ws2.append([d[k] for k in cols2])
    r = ws2.max_row
    ws2.cell(row=r, column=10).number_format = '0.00'
    if d['Multimarca?'] == 'Sim':
        ws2.cell(row=r, column=8).fill = OKF
widths2 = [26, 38, 16, 40, 11, 15, 42, 12, 22, 9]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header(ws2, len(cols2))

# Aba Metodologia
ws3 = wb.create_sheet('Metodologia')
texto = [
    ('CÁLCULO DE MULTIMARCA POR CLIENTE — Ciclo 08/2026', True),
    ('', False),
    ('Fonte: ' + ARQ, False),
    ('Unidades: Matriz Penedo (canal 13707) e Filial Palmeira dos Índios (canal 13706).', False),
    ('', False),
    ('REGRA OFICIAL:', True),
    ('1. A unidade de contagem é o CLIENTE (revendedor), não o pedido.', False),
    ('2. Cliente ATIVO = revendedor com ao menos 1 pedido no ciclo.', False),
    ('3. Cliente MULTIMARCA = comprou 2 ou mais marcas distintas somando TODOS os pedidos do ciclo.', False),
    ('   (marcas guarda-chuva: O Boticário, Eudora, Quem Disse Berenice, O.U.I, AuAmigos)', False),
    ('4. Cada cliente conta UMA vez no ciclo (pedidos repetidos não geram pontos extras).', False),
    ('5. O cliente é creditado ao vendedor que fez o 1º pedido dele no ciclo.', False),
    ('', False),
    ('PONTUAÇÃO:', True),
    ('   - Cliente monomarca .................................. 0 ponto', False),
    ('   - Cliente multimarca (construído ao longo do ciclo) .. 1,00 ponto', False),
    ('   - Cliente multimarca JÁ no 1º pedido (bônus +2%) ..... 1,02 ponto', False),
    ('', False),
    ('% MULTIMARCA = (soma dos pontos / clientes ativos) x 100, limitado a 100%.', False),
    ('Meta institucional: 72%.', False),
]
for i, (t, b) in enumerate(texto, 1):
    cc = ws3.cell(row=i, column=1, value=t)
    if b:
        cc.font = Font(bold=True, size=12 if i == 1 else 11, color='305496' if i == 1 else '000000')
ws3.column_dimensions['A'].width = 95

OUT = 'Multimarca_Clientes_2026-06-18.xlsx'
wb.save(OUT)
print('Planilha gerada:', OUT)
print('Vendedores no resumo:', len([r for r in resumo if r['Código']]))
print('Clientes no detalhe:', len(detalhe))
for un in ordem_un:
    a, p = un_tot[un]
    if a:
        print(f'  {un}: ativos={a} pontos={round(p,1)} => {round(min(100,p/a*100),1)}%')
